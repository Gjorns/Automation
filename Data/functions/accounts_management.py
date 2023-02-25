import os
import json
import openpyxl
import shutil

def check_estimator_id(estimator_id):
    with open('Data/Keno-bi database/estimator.json', 'r') as f:
        estimator_data = json.load(f)
    return estimator_id in estimator_data


def check_account_id(account_id):
    with open('Data/Keno-bi database/account.json', 'r') as f:
        account_data = json.load(f)
    return account_id in account_data

def check_account_name(account_name):
    with open('Data/Keno-bi database/account.json', 'r') as f:
        account_data = json.load(f)
    return any(data.get('account_name', '').lower() == account_name.lower() for data in account_data.values())

def create_new_price_file(estimator_id):
    if not check_estimator_id(estimator_id):
        print(f"Error: Estimator ID {estimator_id} not found")
        return

    with open("Data/Keno-bi database/general.json", "r") as f:
        general_data = json.load(f)
        nacet_path = general_data['nacet']

    while True:
        account_id = input("Enter Account ID: ").strip().upper()
        if check_account_id(account_id):
            print(f"Error: Account ID '{account_id}' already exists")
            continue

        account_name = input("Enter Account name: ").strip().title()
        if check_account_name(account_name):
            print(f"Error: Account name '{account_name}' already exists")
            continue

        print("Select Account Vertical:")
        print("1. New Build")
        print("2. Social Housing")
        print("3. Private & Domestic")
        while True:
            account_vertical = input("(Enter a number): ").strip()
            if account_vertical == '1':
                account_vertical = 'New Build'
                break
            elif account_vertical == '2':
                account_vertical = 'Social Housing'
                break
            elif account_vertical == '3':
                account_vertical = 'Private & Domestic'
                break
            else:
                print("Invalid input, please try again")

        print("Select Account Type:")
        print("1. SLP")
        print("2. MARGIN")
        while True:
            account_type = input("(Enter a number): ").strip()
            if account_type == '1':
                account_type = 'SLP'
                break
            elif account_type == '2':
                account_type = 'MARGIN'
                break
            else:
                print("Invalid input, please try again")

        account_mngr_firstname = input("Enter Account Manager first name: ").strip().capitalize()
        account_mngr_lastname = input("Enter Account Manager last name: ").strip().capitalize()

        while True:
            copper_file = input("Does account have copper file? (Y/N): ").strip().upper()
            if copper_file == 'Y':
                copper_file = True
                break
            elif copper_file == 'N':
                copper_file = False
                break
            else:
                print("Invalid input, please try again")

        account_owner = estimator_id

        account_dir_path = os.path.join(nacet_path, "accounts", account_vertical, account_name)
        account_file_path = os.path.join(account_dir_path, f"{account_id} - {account_name} Master.xlsx")
        os.makedirs(account_dir_path, exist_ok=True)

        price_file_path = os.path.join(general_data['templates'], "Price File.xlsx")
        new_file_path = os.path.join(account_dir_path, f"{account_id} - {account_name} Master.xlsx")
        shutil.copy(price_file_path, new_file_path)

        wb = openpyxl.load_workbook(new_file_path)
        sheet = wb.active

        sheet.unmerge_cells("B1:B2")
        sheet['B1'] = account_name
        sheet.unmerge_cells("B3:B4")
        sheet['B3'] = account_id
        sheet.unmerge_cells("B5:B6")
        sheet['B5'] = account_type

        sheet.merge_cells('B1:B2')
        sheet.merge_cells('B3:B4')
        sheet.merge_cells('B5:B6')

        wb.save(new_file_path)

        with open("Data/Keno-bi database/account.json", "r") as f:
            account_data = json.load(f)

        account_data[account_id] = {
            "account_name": account_name,
            "account_vertical": account_vertical,
            "account_type": account_type,
            "account_mngr_firstname": account_mngr_firstname,
            "account_mngr_lastname": account_mngr_lastname,
            "copper_file": copper_file,
            "account_owner": account_owner,
            "account_file_path": account_file_path
        }

        with open("Data/Keno-bi database/account.json", "w") as f:
            json.dump(account_data, f, indent=4)

        print(f"New account '{account_name}' with ID '{account_id}' has been created!")
        input("Press any key to continue...")
        break

def list_all_accounts(estimator_id):
    with open('Data/Keno-bi database/account.json', 'r') as f:
        account_data = json.load(f)
    for key in account_data:
        owner_id = account_data[key]['account_owner']
        if owner_id == estimator_id:
            account_name = account_data[key]['account_name']
            account_vertical = account_data[key]['account_vertical']
            account_type = account_data[key]['account_type']
            print(f"{key} - {account_name} {account_vertical} {account_type}")
    input("Press Enter to continue...")
    os.system('cls' if os.name == 'nt' else 'clear')