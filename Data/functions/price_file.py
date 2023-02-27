import json
import os
import openpyxl
import math

def pricing_range(current_month):
    if current_month == "January":
        return [7, 10]
    elif current_month == "February":
        return [10, 13]
    elif current_month == "March":
        return [13, 16]
    elif current_month == "April":
        return [16, 19]
    elif current_month == "May":
        return [19, 22]
    elif current_month == "June":
        return [22, 25]
    elif current_month == "July":
        return [25, 28]
    elif current_month == "August":
        return [28, 31]
    elif current_month == "September":
        return [31, 34]
    elif current_month == "October":
        return [34, 37]
    elif current_month == "November":
        return [37, 40]
    elif current_month == "December":
        return [40, 7]
    else:
        return None

def update_additions(estimator_id, current_month):
    # open account.json
    with open("Data\\Keno-bi database\\account.json") as f:
        account_data = json.load(f)

    # open estimator.json
    with open("Data\\Keno-bi database\\estimator.json") as f:
        estimator_data = json.load(f)

        # open additions file using estimator_id looking into estimator.json estimator_path joining with "Price File Additions.xlsx"
        estimator_path = estimator_data[estimator_id]["estimator_path"]
        additions_file = os.path.join(estimator_path, "Price File Additions.xlsx")
        
        try:
            workbook = openpyxl.load_workbook(additions_file)
            worksheet = workbook.active
        except FileNotFoundError:
            print(f"{additions_file} not found. Please check the file path and try again.")
            return
        except Exception as e:
            print(f"An error occurred while opening {additions_file}: {e}")
            return

        # iterate through column A of price file additions, determine how many unique account IDs there are, and store them in a list (we will need to loop it later)
        account_ids = set()
        for row in range(2, worksheet.max_row + 1):
            account_ids.add(worksheet.cell(row, 1).value)

        # Loop should start from first item in the list, and open the account file using the account ID from account.json (account_file_path)
    for account_id in account_ids:
        account_file = account_data.get(account_id, {}).get("account_file_path")
        if account_file is None or not os.path.isfile(account_file):
            print(f"{account_file} not found for account ID {account_id}. Skipping account.")
            continue

        try:
            account_workbook = openpyxl.load_workbook(account_file)
            account_worksheet = account_workbook.active
        except Exception as e:
            print(f"An error occurred while opening {account_file}: {e}. Skipping account ID {account_id}.")
            continue

        # Initialize the price_updates dictionary
        price_updates = {}
        
        # check if the product code in the additions file is in the account file (column A)
        for row in range(2, worksheet.max_row + 1):
            if worksheet.cell(row, 1).value == account_id:
                product_code = worksheet.cell(row, 2).value
                price = worksheet.cell(row, 3).value
                price = math.trunc(price * 1000) / 1000  # truncate to 3 decimal places

                if not any(product_code == account_worksheet.cell(row, 1).value for row in range(12, account_worksheet.max_row + 1)):
                    account_worksheet.append([product_code] + [None] * (account_worksheet.max_column - 1))

                column_range = pricing_range(current_month)
                # if it is, update the price in the account file
                for row in range(12, account_worksheet.max_row + 1):
                    if account_worksheet.cell(row, 1).value == product_code:
                        if product_code not in price_updates:
                            price_updates[product_code] = price
                        else:
                            price_updates[product_code] = price

        # if it is not, add the product code to the account file and update the price for that product code
        for product_code, price in price_updates.items():
            if not any(product_code == account_worksheet.cell(row, 1).value for row in range(12, account_worksheet.max_row + 1)):
                account_worksheet.append([product_code] + [None] * (account_worksheet.max_column - 1))

            for row in range(12, account_worksheet.max_row + 1):
                if account_worksheet.cell(row, 1).value == product_code:
                    column_range = pricing_range(current_month)
                    account_worksheet.cell(row, column_range[0]).value = price
                    account_worksheet.cell(row, column_range[1]).value = price

        # Remove processed rows from the additions file
        for row in range(worksheet.max_row, 1, -1):
            if worksheet.cell(row, 1).value == account_id:
                product_code = worksheet.cell(row, 2).value
                if any(product_code == account_worksheet.cell(row, 1).value for row in range(12, account_worksheet.max_row + 1)):
                    worksheet.delete_rows(row, amount=1)

        account_workbook.save(account_file)
        account_workbook.close()

    workbook.save(additions_file)
    workbook.close()

    print("Additions updated. Press enter to continue...")
    input()